import csv
import json
from tempfile import NamedTemporaryFile
import collections
import threading
from functools import reduce
import operator
from operator import itemgetter
from datetime import (
    datetime,
    date,
    timedelta,
)
import logging
from calendar import (
    monthrange,
    monthcalendar
)

from django.core.cache import cache
from django.db.models import (
    Max,
    Min,
    Count,
    Q,
)
from django.db.models.functions import (
    Trunc,
    ExtractDay,
)
from django.db import connection
from django.utils import timezone
from django.utils.timezone import make_aware

from openpyxl import Workbook
import requests

from taicat.models import (
    Project,
    Image,
    StudyArea,
    Deployment,
    DeploymentJournal,
    DeploymentStat,
    ProjectMember,
    Contact,
    Organization,
    Species,
    ProjectSpecies,
    ProjectStat,
    HomePageStat,
    DeletedImage,
    Calculation,
    timezone_utc_to_tw,
    timezone_tw_to_utc,
)

import geopandas as gpd


# WIP
def display_working_day_in_calendar_html(year, month, working_day):
    month_cal = monthcalendar(year, month)
    s = '| 一 | 二 | 三 | 四 | 五 | 六 | 日 |'
    
    for week in month_cal:
        for d in week:
            s += f'{d}'
    '''
    print (month_cal)
    s = '<table class="table"><tr><th>一</th><th>二</th><th>三</th><th>四</th><th>五</th><th>六</th><th>日</th></tr>'
    for week in month_cal:
        s += '<tr>'
        for d in week:
            #if d > 0:
            #    if working_day[d-1]:
            #        pass
                    #s
            #else:
            #    pass
            #    #res_week.append('0')
            s += '<td>{}</td>'.format(d)
        s += '</tr>'
    s += '</table>'
    '''
    return s

def find_deployment_working_day(year, month, dep_id=''):
    num_month = monthrange(year, month)[1]
    month_start = make_aware(datetime(year, month, 1))
    month_end = make_aware(datetime(year, month, num_month))
    month_stat = [0] * num_month

    query = DeploymentJournal.objects.filter(
        is_effective=True,
        deployment_id=dep_id,
        working_start__lte=month_end,
        working_end__gte=month_start).order_by('working_start')

    ret = []
    for i in query.all():
        # updated
        #print ('-------')
        #print (i.working_start.toordinal(), i.working_end.toordinal(), dt1.toordinal(), dt2.toordinal())
        month_stat_part = [0] * num_month
        overlap_range = [max(i.working_start, month_start), min(i.working_end, month_end)]
        gap_days = (overlap_range[0]-month_start).days
        duration_days = (overlap_range[1]-overlap_range[0]).days+1
        for index, stat in enumerate(month_stat):
            if index >= gap_days and index < gap_days + duration_days:
                    month_stat[index] = 1
                    month_stat_part[index] = 1
            #print(month_stat_part)
        ret.append([i.working_start.strftime('%Y-%m-%d'), i.working_end.strftime('%Y-%m-%d')])
    #print('fin',dep_id, year, month,  month_stat)
    return month_stat, ret


def get_species_list(force_update=False):
    CACHE_KEY = 'species_list'
    species_list = cache.get(CACHE_KEY)
    if not force_update and species_list:
        return species_list
    else:
        species_list = count_all_species_list()
        #print('save', species_list)
        cache.set(CACHE_KEY, species_list, 86400) # 1d
        return species_list

def count_all_species_list():
    all_species_list = []
    ret = {}
    for p in Project.objects.all():
        img_list = Image.objects.values_list('annotation', flat=True).filter(project_id=p).all()
        project_species_list = []
        for alist in img_list:
            for a in alist:
                try:
                    if sp := a.get('species', ''):
                        project_species_list.append(sp)
                        all_species_list.append(sp)
                except:
                    #print ('annotation load error')
                    pass
        counter = collections.Counter(project_species_list)
        counter_dict = dict(counter)
        project_species_list = sorted(counter_dict.items(), key=itemgetter(1), reverse=True)
        #print(p, p.id, project_species_list, len(img_list))
        ret[p.id] = project_species_list
    counter_all = collections.Counter(all_species_list)
    counter_dict_all = dict(counter_all)
    all_species_list = sorted(counter_dict_all.items(), key=itemgetter(1), reverse=True)
    ret['all'] = all_species_list

    return ret


def calc_from_cache(filter_args, calc_args):
    # print (filter_args, calc_args)
    deps = filter_args.get('deployments')
    species = filter_args.get('species')
    start_dt = None
    end_dt = None
    if start_date := filter_args.get('startDate', ''):
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')

    if end_date := filter_args.get('endDate', ''):
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')

    image_interval = calc_args.get('imageInterval')
    event_interval = calc_args.get('eventInterval')
    #dt = datetime.strptime(start_date, '%Y-%m-%d')
    results = {}
    for sp in species:
        results[sp] = []
        for did in deps:
            dep = Deployment.objects.get(pk=did)
            #stats = dep.project.get_or_count_stats()
            #w = stats['working__range']
            #start_dt = datetime.fromtimestamp(w[0]).year
            #end_dt = datetime.fromtimestamp(w[1]).year
            # init from stats
            if c := dep.get_calc_cache():
                for year in range(start_dt.year, end_dt.year+1):
                    for month in range(1, 13):
                        dt = datetime(year=year, month=month, day=1)
                        if dt >= start_dt and dt <= end_dt:
                            #print (year, month)
                            for i in c:
                                key = f'{did}/{year}/{month}/{sp}/{image_interval}/{event_interval}'
                                if i[0] == key:
                                    #print (key, stats['years'][str(year)])
                                    results[sp].append(i)


    return results

def calculated_data(filter_args, calc_args):
    # print (filter_args, calc_args)
    deps = filter_args.get('deployments')
    species = filter_args.get('species')
    start_dt = None
    end_dt = None

    if start_date := filter_args.get('startDate', ''):
        start_dt =timezone_tw_to_utc(datetime.strptime(start_date, '%Y-%m-%d'))
        # ex: 2022-11-30 16:00:00:00
        start_dt = make_aware(start_dt)

    if end_date := filter_args.get('endDate', ''):
        end_dt = timezone_tw_to_utc(datetime.strptime(end_date, '%Y-%m-%d')+timedelta(days=1))
        # ex: 2022-12-31 16:00:00:00
        end_dt = make_aware(end_dt) 

    image_interval = calc_args.get('imageInterval')
    event_interval = calc_args.get('eventInterval')
    #dt = datetime.strptime(start_date, '%Y-%m-%d')
    results = {}

    query = Deployment.objects.filter(id__in=deps)

    if value := filter_args.get('altitude'):
        if op := filter_args.get('altitudeOperator'):
            val_list = value.split('-')
            v1 = int(val_list[0])
            v2 = None
            if len(val_list) >= 2:
                v2 = int(val_list[1])
            if v1:
                if op == 'eq':
                    query = query.filter(altitude=v1)
                elif op == 'gt':
                    query = query.filter(altitude__gte=v1)
                elif op == 'lt':
                    query = query.filter(altitude__lte=v1)
            if v2 and op == 'range':
                query = query.filter(altitude__gte=v1, altitude__lte=v2)

    if values := filter_args.get('counties'):
        q_list = []
        for x in values:
            q_list.append(Q(county__icontains=x['parametername']))

        query = query.filter(reduce(operator.or_, q_list))

    if values := filter_args.get('protectedareas'):
        q_list = []
        for x in values:
            # 只有一筆就用 eq, 多筆就加上前後 ","
            q_list.append(Q(protectedarea__icontains=x['parametername']+','))
            q_list.append(Q(protectedarea__icontains=','+x['parametername']))
            q_list.append(Q(protectedarea=x['parametername']))

        query = query.filter(reduce(operator.or_, q_list))

    deployment_list = query.values('id', 'name', 'project__name', 'study_area__name').all()
    #print(deployment_list)
    for sp in species:
        results[sp] = []
        cal_query = Calculation.objects.filter(
            datetime_from__gte=start_dt,
            datetime_to__lte=end_dt,
            image_interval=image_interval,
            event_interval=event_interval,
            species=sp
        )
        if deps:
            dep_ids = [x['id'] for x in deployment_list]
            cal_query = cal_query.filter(deployment_id__in=dep_ids)

        #print (cal_query.query, start_dt, end_dt)
        if res := cal_query.all():
            for cal in res:
                # print (cal.datetime_from, cal.datetime_to, cal.id)
                #results[sp].append(cal.data)
                dep = cal.deployment
                results[sp].append({
                    'project': dep.project.name,
                    'studyarea': dep.study_area.name,
                    'name': dep.name,
                    'year': cal.datetime_to.year,
                    'month': cal.datetime_to.month,
                    'calc': cal.data
                })

    return results

def calc(query, calc_data, query_start, query_end):
    #print('query', calc_data)

    #agg = query.aggregate(Max('datetime'), Min('datetime'))

    results = {} # by species/deployment/year/month
    # group by species
    species_list = query.values('species').annotate(count=Count('species')).order_by()
    # group by deployment
    deployment_list = query.values('deployment_id').annotate(count=Count('deployment')).order_by()
    proj_info = {}
    for d in deployment_list:
        if deployment := Deployment.objects.get(pk=d['deployment_id']):
            pid = deployment.project_id
            if pid not in proj_info:
                # cached = deployment.project.get_or_count_stats()
                # w = cached['working__range']
                proj_info[pid] ={
                    'deployments': [],
                    #'cached': deployment.project.get_or_count_stats()
                    #'working_range': [datetime.fraomtimestamp(w[0]), datetime.fromtimestamp(w[1])]
                    'working_range': [query_start, query_end]
                }
            proj_info[pid]['deployments'].append(deployment)

    # print(deployment_list, query.query)
    # print(proj_info)
    for species in species_list:
        species_name = species['species']
        results[species_name] = []
        for pid, info in proj_info.items():
            for dep in info['deployments']:
                for year in range(info['working_range'][0].year, info['working_range'][1].year+1):
                    for month in range(1, 13):
                        if year == info['working_range'][0].year and month <= info['working_range'][0].month -1:
                            # first year, check start month
                            # print(pid, year, month, 'start pass')
                            pass
                        elif year == info['working_range'][1].year and month > info['working_range'][1].month:
                            # print(pid, year, month, 'end pass')
                            pass
                        else:
                            # print (pid, dep.project.name, year, month, 'normal')
                            res = deployment.calculate(year, month, species_name,
                                                       int(calc_data['imageInterval']),
                                                       int(calc_data['eventInterval']))
                            results[species_name].append({
                                'project': dep.project.name,
                                'studyarea': dep.study_area.name,
                                'name': dep.name,
                                'year': year,
                                'month': month,
                                'calc': res,
                            })
    return results

def calc_output_file(results, file_format, filter_str, calc_str, target_file=''):
    '''
    filter_str, calc_str for display query condition
    '''
    if file_format == 'csv':
        '''csv 多物種會全部放在一個大表
        not test (2023.05.05)
        '''
        #with tempfile.TemporaryFile('w+t') as fp:
        with NamedTemporaryFile('w+t') as tmp:
            tmp.write('filters:'+ filter_str + 'calc:' + calc_str + '\n')
            tmp.write('==='+'\n')
            tmp.write('year,month,物種,days in month,相機位置,相機工作時數,有效照片數,目擊事件數,OI3,捕獲回合比例,存缺,'+','.join(f'活動機率day{day}' for day in range(1, 32))+'\n')
            for sp in results:
                for y in results[sp]:
                    for m, value in enumerate(results[sp][y]):
                        row = []
                        for x in value:
                            if x != 'apoa':
                                row.append(str(value[x]))
                            else:
                                if value[x] != None:
                                    for day in value[x]:
                                        hours = '|'.join([str(d) for d in day])
                                        row.append(hours)

                        #print (row)
                        tmp.write(','.join(row) + '\n')

            tmp.seek(0)
            return tmp.read()

    elif file_format == 'excel':
        wb = Workbook()
        ws = wb.active
        query_condition = 'filters:'+ filter_str + 'calc:' + calc_str
        calcData = json.loads(calc_str)
        ws.cell(row=1, column=1, value=query_condition)
        sheets = []
        sheet_index = 0
        for sp in results:
            row_index = 1
            sheets.append(wb.create_sheet(title=sp))
            if calcData.get('calcType') == 'basic-oi':
                header_str = '計劃,樣區,相機位置,年,月,物種,相機工作時數,有效照片數,目擊事件數,OI1,OI2,OI3,'+','.join(f'相機工作天{day}' for day in range(1, 32))  # ,捕獲回合比例,存缺,'+','.join(f'活動機率day{day}' for day in range(1, 32))
                header_list = header_str.split(',')
                for h, v in enumerate(header_list):
                    sheets[sheet_index].cell(row=1, column=h+1, value=v)

                for i in results[sp]:
                    row_index += 1
                    sheets[sheet_index].cell(row=row_index, column=1, value=i['project'])
                    sheets[sheet_index].cell(row=row_index, column=2, value=i['studyarea'])
                    sheets[sheet_index].cell(row=row_index, column=3, value=i['name'])
                    sheets[sheet_index].cell(row=row_index, column=4, value=i['year'])
                    sheets[sheet_index].cell(row=row_index, column=5, value=i['month'])
                    sheets[sheet_index].cell(row=row_index, column=6, value=sp)
                    sheets[sheet_index].cell(row=row_index, column=7, value=sum(i['calc'][0])*24 if i['calc'][0] else '')
                    sheets[sheet_index].cell(row=row_index, column=8, value=i['calc'][1])
                    sheets[sheet_index].cell(row=row_index, column=9, value=i['calc'][2])
                    sheets[sheet_index].cell(row=row_index, column=10, value=i['calc'][3])
                    sheets[sheet_index].cell(row=row_index, column=11, value='')
                    sheets[sheet_index].cell(row=row_index, column=12, value=i['calc'][5])
                    for day_index, working in enumerate(i['calc'][0]):
                        sheets[sheet_index].cell(row=row_index, column=13+day_index, value=working)

            elif calcData.get('calcType') == 'pod':
                header_str = '計劃,樣區,相機位置,年,月,物種,拍到天數,相機工作天數,POD,' + ','.join(f'偵測到/未偵測到{day_index+1}' for day_index in range(0, 31))
                header_list = header_str.split(',')
                for h, v in enumerate(header_list):
                    sheets[sheet_index].cell(row=1, column=h+1, value=v)

                for i in results[sp]:
                    row_index += 1
                    count = 0
                    for d in i['calc'][7]:
                        if d[0]:
                            count += 1
                    sum_working_days = sum(i['calc'][0])
                    sheets[sheet_index].cell(row=row_index, column=1, value=i['project'])
                    sheets[sheet_index].cell(row=row_index, column=2, value=i['studyarea'])
                    sheets[sheet_index].cell(row=row_index, column=3, value=i['name'])
                    sheets[sheet_index].cell(row=row_index, column=4, value=i['year'])
                    sheets[sheet_index].cell(row=row_index, column=5, value=i['month'])
                    sheets[sheet_index].cell(row=row_index, column=6, value=sp)
                    sheets[sheet_index].cell(row=row_index, column=7, value=count)
                    sheets[sheet_index].cell(row=row_index, column=8, value=sum_working_days)
                    sheets[sheet_index].cell(row=row_index, column=9, value=count*1.0/sum_working_days if sum_working_days > 0 else 'N/A')
                    for day_index, d in enumerate(i['calc'][7]):
                        sheets[sheet_index].cell(row=row_index, column=10 + day_index, value=d[0])

            elif calcData.get('calcType') == 'apoa':
                header_str = '計劃,樣區,相機位置,物種,date,' + ','.join(f'{hour:02}' for hour in range(0, 24))  # ,捕獲回合比例,存缺,'+','.join(f'活動機率day{day}' for day in range(1, 32))
                header_list = header_str.split(',')
                for h, v in enumerate(header_list):
                    sheets[sheet_index].cell(row=1, column=h+1, value=v)

                for i in results[sp]:
                    for day_index, hours in enumerate(i['calc'][7]):
                        row_index += 1
                        sheets[sheet_index].cell(row=row_index, column=1, value=i['project'])
                        sheets[sheet_index].cell(row=row_index, column=2, value=i['studyarea'])
                        sheets[sheet_index].cell(row=row_index, column=3, value=i['name'])
                        sheets[sheet_index].cell(row=row_index, column=4, value=sp)
                        sheets[sheet_index].cell(row=row_index, column=5, value=f"{i['year']}-{i['month']}-{day_index+1}")
                        for hour_index, hour in enumerate(hours[1]):
                            sheets[sheet_index].cell(row=row_index, column=6+hour_index, value=hour)

            sheet_index += 1

            if target_file:
                wb.save(target_file)
            #wb.save(tmp.name)
            #tmp.seek(0)
            #return tmp.read()
            return target_file

def calc_output2(results, file_format, filter_str, calc_str):
    '''
    filter_str, calc_str for display query condition
    '''
    if file_format == 'csv':
        '''csv 多物種會全部放在一個大表
        '''
        #with tempfile.TemporaryFile('w+t') as fp:
        with NamedTemporaryFile('w+t') as tmp:
            tmp.write('filters:'+ filter_str + 'calc:' + calc_str + '\n')
            tmp.write('==='+'\n')
            tmp.write('year,month,物種,days in month,相機位置,相機工作時數,有效照片數,目擊事件數,OI3,捕獲回合比例,存缺,'+','.join(f'活動機率day{day}' for day in range(1, 32))+'\n')
            for sp in results:
                for y in results[sp]:
                    for m, value in enumerate(results[sp][y]):
                        row = []
                        for x in value:
                            if x != 'apoa':
                                row.append(str(value[x]))
                            else:
                                if value[x] != None:
                                    for day in value[x]:
                                        hours = '|'.join([str(d) for d in day])
                                        row.append(hours)

                        #print (row)
                        tmp.write(','.join(row) + '\n')

            tmp.seek(0)
            return tmp.read()

    elif file_format == 'excel':
        with NamedTemporaryFile() as tmp:
            wb = Workbook()
            ws = wb.active
            query_condition = 'filters:'+ filter_str + 'calc:' + calc_str
            ws.cell(row=1, column=1, value=query_condition)
            sheets = []
            sheet_index = 0
            for sp in results:
                row_index = 1
                sheets.append(wb.create_sheet(title=sp))
                header_str = 'year,month,物種,相機位置,相機工作時數,有效照片數,目擊事件數,OI3'  # ,捕獲回合比例,存缺,' +','.join(f'活動機率day{day}' for day in range(1, 32))
                header_list = header_str.split(',')
                for h, v in enumerate(header_list):
                    sheets[sheet_index].cell(row=1, column=h+1, value=v)

                for i in results[sp]:
                    klist = i[0].split('/')
                    row_index += 1
                    sheets[sheet_index].cell(row=row_index, column=1, value=klist[1])
                    sheets[sheet_index].cell(row=row_index, column=2, value=klist[2])
                    sheets[sheet_index].cell(row=row_index, column=3, value=klist[3])
                    sheets[sheet_index].cell(row=row_index, column=4, value=i[1])
                    sheets[sheet_index].cell(row=row_index, column=5, value=sum(i[2])*24 if i[2] else '')
                    sheets[sheet_index].cell(row=row_index, column=6, value=i[3][0])
                    sheets[sheet_index].cell(row=row_index, column=7, value=i[3][1])
                    sheets[sheet_index].cell(row=row_index, column=8, value=i[3][4])

                sheet_index += 1

            wb.save(tmp.name)
            tmp.seek(0)
            return tmp.read()

def get_my_project_list(member_id, project_list=[]):
    if Contact.objects.filter(id=member_id, is_system_admin=True).exists():
        project_list += list(Project.objects.all().values_list('id', flat=True))
    else:
        # 1. select from project_member table
        with connection.cursor() as cursor:
            query = "SELECT project_id FROM taicat_projectmember where member_id ={}"
            cursor.execute(query.format(member_id))
            temp = cursor.fetchall()
            for i in temp:
                if i[0]:
                    project_list += [i[0]]
        # 2. check if the user is organization admin
        if_organization_admin = Contact.objects.filter(id=member_id, is_organization_admin=True)
        if if_organization_admin:
            organization_id = if_organization_admin.values('organization').first()['organization']
            temp = Organization.objects.filter(id=organization_id).values('projects')
            for i in temp:
                project_list += [i['projects']]
    return project_list


def get_project_member(project_id):
    member_list = []
    members = [m.member_id for m in ProjectMember.objects.filter(project_id=project_id)]
    organization_id = Organization.objects.filter(projects=project_id).values('id')
    for i in organization_id:
        members += [c.id for c in Contact.objects.filter(organization=i['id'], is_organization_admin=True)]
    for m in members: # 排除重複
        if m not in member_list:
            member_list += [m]
    return member_list

def get_studyarea_member(project_id,studyarea_id):
    member_list = []
    members = [m.member_id for m in StudyArea.objects.get(id=studyarea_id).projectmember_set.all()]
    organization_id = Organization.objects.filter(projects=project_id).values('id')

    for i in organization_id:
        members += [c.id for c in Contact.objects.filter(organization=i['id'], is_organization_admin=True)]
    for m in members: # 排除重複
        if m not in member_list:
            member_list += [m]
    return member_list


def get_none_studyarea_project_member(project_id,role):
    uploader = [m.member_id for m in ProjectMember.objects.filter(project_id=project_id,role__in=role,pmstudyarea__isnull=True)]

    return uploader

def sanitize_date(input):
    if len(input) != 8:
        return ''

    year = input[0:4]
    month = input[4:6]
    day = input[6:8]
    return f'{year}-{month}-{day}'

def set_deployment_journal(data, deployment):
    # if data.get('trip_start') and data.get('trip_end') and data.get('folder_name') and data.get('source_id'):  # 不要判斷了

    is_new = None
    dj_exist = DeploymentJournal.objects.filter(
        deployment=deployment,
        folder_name=data['folder_name'],
        local_source_id=data['source_id']).first()

    trip_start = sanitize_date(data['trip_start'])
    trip_end = sanitize_date(data['trip_end'])

    if dj_exist:
        is_new = False
        is_modified = False
        if data.get('trip_start') and data.get('trip_end'):
            if trip_start and trip_end:
                dj_exist.working_start = trip_start
                dj_exist.working_end = trip_end
                is_modified = True

        if is_modified:
            dj_exist.last_updated = timezone.now()
            dj_exist.save()

        deployment_journal_id = dj_exist.id

    else:
        dj_new = DeploymentJournal(
            deployment_id=deployment.id,
            project=deployment.project,
            studyarea=deployment.study_area,
            is_effective=False,
            folder_name=data['folder_name'],
            local_source_id=data['source_id'],
            last_updated=timezone.now()
        )

        is_new = True

        if trip_start:
            dj_new.working_start = trip_start
        if trip_end:
            dj_new.working_end = trip_end

        dj_new.save()

        deployment_journal_id = dj_new.id

    obj = dj_new if is_new is True else dj_exist

    # 有時間才算有效
    if trip_start and trip_end:
        obj.is_effective = True
        obj.save()

    # update cache
    #task = threading.Thread(target=obj.project.get_or_count_stats, args=(True,))
    #task.start()

    return obj

def clone_image(obj):
    new_img = Image(
        deployment=obj.deployment,
        project=obj.project,
        studyarea=obj.studyarea,
        file_url=obj.file_url,
        filename=obj.filename,
        datetime=obj.datetime,
        count=obj.count,
        image_hash=obj.image_hash,
        image_uuid=obj.image_uuid,
        has_storage=obj.has_storage,
        folder_name=obj.folder_name,
        specific_bucket=obj.specific_bucket,
        deployment_journal=obj.deployment_journal
    )
    new_img.save()
    return new_img

def set_image_annotation(image_obj):
    '''extract annotation (json field) to seperate field, ex: species, life_stage...
    '''
    def update_annotation_field(obj, values):
        # (annotation_field, model field)
        field_map = {
            'species': 'species',
            'lifestage': 'life_stage', # 該死的 _
            'sex': 'sex',
            'antler': 'antler',
            'remark': 'remarks', # 該死的 s
            'animal_id': 'animal_id',
        }
        for field in field_map:
            if v := values.get(field):
                setattr(obj, field_map[field], v)
        obj.last_updated = timezone.now()

    if annotation := image_obj.annotation:
        related_images = Image.objects.filter(image_uuid=image_obj.image_uuid).order_by('annotation_seq').all()
        num_related = len(related_images)
        num_annotation = len(annotation)
        # print (num_related, num_annotation, flush=True)
        to_delete = []
        for i in range(0, max(num_annotation, num_related)):
            if i == 0:  # original image
                update_annotation_field(image_obj, annotation[0])
                image_obj.save()
            elif num_related - num_annotation <=0:  # new annotation (cloned image)
                if i < num_related:
                    update_annotation_field(related_images[i], annotation[i])
                    related_images[i].save()
                else:
                    cloned = clone_image(image_obj)
                    update_annotation_field(cloned, annotation[i])
                    cloned.annotation_seq = i
                    cloned.save()
            elif num_related - num_annotation > 0: # delete cloned image
                if i < num_annotation:
                    update_annotation_field(related_images[i], annotation[i])
                    related_images[i].save()
                else:
                    to_delete = [x.id for x in related_images if x.annotation_seq > 0]
                    break

        if len(to_delete) > 0:
            #print(to_delete, flush=True)
            delete_image_by_ids(to_delete, image_obj.project_id)
            # 全部刪掉再重建
            for i, a in enumerate(image_obj.annotation):
                if i > 0:
                    cloned = clone_image(image_obj)
                    update_annotation_field(cloned, annotation[i])
                    cloned.annotation_seq = i
                    cloned.save()


def delete_image_by_ids(image_list=[], pk=None):
    # mode
    mode = Project.objects.filter(id=pk).first().mode
    now = timezone.now()
    image_objects = Image.objects.filter(id__in=image_list)
    # species的資料先用id抓回來計算再扣掉
    query = image_objects.values('species').annotate(total=Count('species')).order_by('-total')
    for q in query:
        if mode == 'official':
            # taicat_species
            if sp := Species.objects.filter(name=q['species']).first():
                if sp.count == q['total']:
                    sp.delete()
                else:
                    sp.count -= q['total']
                    sp.last_updated = now
                    sp.save()
        # taicat_projectspecies
        if p_sp := ProjectSpecies.objects.filter(name=q['species'], project_id=pk).first():
            if p_sp.count == q['total']:
                p_sp.delete()
            else:
                p_sp.count -= q['total']
                p_sp.last_updated = now
                p_sp.save()

    if ProjectStat.objects.filter(project_id=pk).exists():
        p = ProjectStat.objects.get(project_id=pk)
        p.num_data -= image_objects.count()
        p.last_updated = now
        p.save()

    if mode == 'official':
        year = image_objects.aggregate(Min('datetime'))['datetime__min'].strftime("%Y")
        home = HomePageStat.objects.filter(year__gte=year)
        for h in home:
            h.count -= image_objects.count()
            h.last_updated = now
            h.save()

    # move deleted image to DeletedImage table
    image_dict = image_objects.values()
    for d in image_dict:
        di = DeletedImage(**d)
        di.save()
    Image.objects.filter(id__in=image_list).delete()

    species = ProjectSpecies.objects.filter(project_id=pk).order_by('count').values('count', 'name')
    return list(species)


def half_year_ago(year, month):
    '''前一個月的前半年
    '''
    month_list = list(range(1, 13))
    begin_year = year
    begin_month = None
    end_year = year
    end_month = month_list[month-2]
    if month < 2:
        end_year = year - 1

    if month - 8 >= 0:
        begin_month = month - 7
    else:
        begin_month = month_list[month-8]
        begin_year = year-1

    return [
        datetime.strptime(f'{begin_year}-{begin_month}-01 00:00:00', "%Y-%m-%d %H:%M:%S"),
        datetime.strptime(f'{end_year}-{end_month}-01 00:00:00', "%Y-%m-%d %H:%M:%S")
    ]


def save_calculation(species_list, year, month, deployment):
    print('save_calcultaion: {} {} {} {}'.format(species_list, year, month, deployment.id))
    for sp in species_list:
        # print('sp', sp, datetime_from.year, datetime_from.month)
        species = sp.strip()
        for img_int in [30, 60]:
            for e_int in [2, 5, 10, 30, 60]:
                result = deployment.calculate(year, month, species, img_int, e_int, to_save=True)

def apply_search_filter(filter_dict={}):
    query = Image.objects.filter()
    project_ids = []

    if value := filter_dict.get('keyword'):
        rows = Project.objects.values_list('id', flat=True).filter(keyword__icontains=value)
        project_ids = list(rows)
        if len(project_ids) > 0:
            query = query.filter(project_id__in=project_ids)
        else:
            query = query.filter(project_id__in=[9999]) # 關鍵字沒有就都不要搜到
        #if values := filter_dict.get('projects'):
        #        project_ids = values

    sp_values = []
    if values := filter_dict.get('species'):
        sp_values += values
    if value := filter_dict.get('speciesText'):
        sp_values += [value]

    if value := filter_dict.get('startDate'):
        dt = make_aware(datetime.strptime(value, '%Y-%m-%d'))
        query_start = timezone_utc_to_tw(dt)
        query = query.filter(datetime__gte=dt)
    if value := filter_dict.get('endDate'):
        dt = make_aware(datetime.strptime(value, '%Y-%m-%d'))
        query_end = timezone_utc_to_tw(dt)
        query = query.filter(datetime__lte=dt)
    if values := filter_dict.get('deployments'):
        query = query.filter(deployment_id__in=values)
        #if len(project_ids):
        #    query = query.filter(Q(deployment_id__in=values) | Q(project_id__in=project_ids))
        # else:
        #    query = query.filter(deployment_id__in=values)
    elif values := filter_dict.get('studyareas'):
        query = query.filter(studyarea_id__in=values)

    if value := filter_dict.get('altitude'):
        if op := filter_dict.get('altitudeOperator'):
            val_list = value.split('-')
            v1 = int(val_list[0])
            v2 = None
            if len(val_list) >= 2:
                v2 = int(val_list[1])
            if v1:
                if op == 'eq':
                    query = query.filter(deployment__altitude=v1)
                elif op == 'gt':
                    query = query.filter(deployment__altitude__gte=v1)
                elif op == 'lt':
                    query = query.filter(deployment__altitude__lte=v1)
            if v2 and op == 'range':
                query = query.filter(deployment__altitude__gte=v1, deployment__altitude__lte=v2)

    if values := filter_dict.get('counties'):
        q_list = []
        for x in values:
            q_list.append(Q(deployment__county__icontains=x['parametername']))

        query = query.filter(reduce(operator.or_, q_list))

    if values := filter_dict.get('protectedareas'):
        q_list = []
        for x in values:
            # 只有一筆就用 eq, 多筆就加上前後 ","
            q_list.append(Q(deployment__protectedarea__icontains=x['parametername']+','))
            q_list.append(Q(deployment__protectedarea__icontains=','+x['parametername']))
            q_list.append(Q(deployment__protectedarea=x['parametername']))
        query = query.filter(reduce(operator.or_, q_list))

    if len(sp_values) > 0:
        query = query.filter(species__in=sp_values)

    return query

def humen_readable_filter(filter_dict):
    data = []
    if species := filter_dict.get('species'):
        data.append(f"物種：{species}")
    start = filter_dict.get('startDate', '')
    end= filter_dict.get('endDate', '')
    if start or end:
        data.append(f'時間：{start} ~ {end}')
    if counties := filter_dict.get('counties'):
        a = [x['name'] for x in counties]
        data.append(f"縣市：{'|'.join(a)}")
    if protectedareas := filter_dict.get('protectedareas'):
        a = [x['name'] for x in protectedareas]
        data.append(f"保護留區：{'|'.join(a)}")
    if keyword := filter_dict.get('keyword'):
        data.append(f'計畫關鍵字：{keyword}')
    if alt := filter_dict.get('altitude'):
        alt_map = {'eq': '等於', 'gt': '>', 'lt': '<'}
        data.append(f"海拔：{alt_map[filter_dict['altitudeOperator']]}{alt}")

    project_dict = {}
    if verbose := filter_dict.get('verbose'):
        if projects := verbose.get('projects'):
            for p in projects:
                project_dict[p['project']['name']] = []
                if studyareas := p.get('studyareas'):
                    for sa in studyareas:
                        project_dict[p['project']['name']].append(f"樣區：{sa['name']}")
                if deployments := p.get('deployments'):
                    for d in deployments:
                        project_dict[p['project']['name']].append(f"相機位置：{d['name']}")
    for k, v in project_dict.items():
        detail = ' AND '.join(v)
        data.append(f'計畫名稱：{k}, {detail}')

    return ', '.join(data)
